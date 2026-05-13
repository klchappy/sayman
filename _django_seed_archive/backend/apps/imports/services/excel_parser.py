"""
Excel parser temeli — openpyxl ile sheet okuma + draft yaratma.

Faz 3 kapsamı:
- Workbook → sheet listesi
- Sheet → satır JSON
- Boş satır algılama (SKIP)
- Header tahmini (ilk dolu satır default header)
- Domain commit yok — yalnız ImportDraftRecord oluşur.
"""
from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from typing import Any, Iterable, Iterator

import openpyxl
from django.conf import settings


def _normalize_cell_value(value: Any) -> Any:
    """openpyxl cell value'sunu JSON-safe hale getir."""
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def list_sheets(file_obj) -> list[str]:
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def iter_sheet_rows(
    file_obj,
    sheet_name: str | None = None,
    *,
    max_rows: int | None = None,
) -> Iterator[tuple[int, list[Any]]]:
    """
    Bir sheet'i satır satır okur. (row_number, [cell_value, ...]) yield eder.
    row_number 1-based (Excel uyumlu).

    `max_rows`: settings.IMPORT_PREVIEW_ROWS_LIMIT'i aşamaz.
    """
    limit = settings.IMPORT_PREVIEW_ROWS_LIMIT
    if max_rows is not None:
        limit = min(limit, max_rows)

    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i > limit:
                break
            yield i, [_normalize_cell_value(c) for c in row]
    finally:
        wb.close()


def is_empty_row(values: Iterable[Any]) -> bool:
    return all(v is None or (isinstance(v, str) and v.strip() == "") for v in values)


def detect_header(rows: list[tuple[int, list[Any]]]) -> tuple[int | None, list[str] | None]:
    """
    İlk dolu satırı header olarak alır (basit heuristic).

    Returns: (header_row_number, [normalized_column_name, ...])
    """
    for row_no, vals in rows:
        if is_empty_row(vals):
            continue
        # En az 2 dolu hücre varsa header kabul et
        non_empty = [v for v in vals if v not in (None, "")]
        if len(non_empty) >= 2:
            headers = [
                (str(v).strip() if v is not None else f"col_{i+1}").lower().replace(" ", "_")[:64]
                for i, v in enumerate(vals)
            ]
            return row_no, headers
    return None, None


def parse_workbook_to_drafts(
    file_obj,
    *,
    batch,
    document,
    sheet_name: str | None = None,
    max_rows: int | None = None,
):
    """
    Workbook → ImportSourceFile + ImportDraftRecord'lar.

    Domain commit yok. Tüm satırlar `validation_status=MANUAL_REVIEW`
    ve `suggested_action=MANUAL_REVIEW` olarak yaratılır (Faz 4'te
    modül-spesifik validator'lar bu durumları yeniden hesaplayacak).

    Returns: dict {sheet_count, total_rows, draft_count, skipped_empty}
    """
    from apps.imports.models import (
        DraftSuggestedAction,
        ImportDraftRecord,
        ImportLog,
        ImportLogLevel,
        ImportSourceFile,
        ValidationStatus,
    )

    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    try:
        sheets = [sheet_name] if sheet_name else wb.sheetnames
        total_rows = 0
        draft_count = 0
        skipped_empty = 0

        for s_name in sheets:
            ws = wb[s_name]
            # Read all rows (within limit)
            rows = list(iter_sheet_rows(file_obj, s_name, max_rows=max_rows))
            # iter_sheet_rows iterator iki kez açıyor olabilir; tekrar oku
            # Alternatif: doğrudan ws.iter_rows kullanalım:
            ws_rows: list[tuple[int, list[Any]]] = []
            for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if max_rows and i > max_rows:
                    break
                ws_rows.append((i, [_normalize_cell_value(c) for c in row]))

            header_row, headers = detect_header(ws_rows)

            source_file = ImportSourceFile.objects.create(
                batch=batch,
                document=document,
                sheet_name=s_name,
                file_role="primary",
                row_count=len(ws_rows),
                status="PARSED",
                metadata={
                    "header_row": header_row,
                    "headers": headers or [],
                },
            )

            for row_no, vals in ws_rows:
                if is_empty_row(vals):
                    skipped_empty += 1
                    continue

                if header_row and row_no == header_row:
                    # Header satırını draft olarak yazma; metadata'da var.
                    continue

                raw = {
                    "row": row_no,
                    "sheet": s_name,
                    "values": vals,
                }
                if headers:
                    # headers ile eşleşmiş dict — raw_data zenginleştirme
                    paired = {}
                    for col_i, h in enumerate(headers):
                        paired[h] = vals[col_i] if col_i < len(vals) else None
                    raw["paired"] = paired

                title_parts = [str(v) for v in vals[:3] if v not in (None, "")]
                display = " · ".join(title_parts)[:255]

                ImportDraftRecord.objects.create(
                    batch=batch,
                    source_file=source_file,
                    source_row_number=row_no,
                    suggested_action=DraftSuggestedAction.MANUAL_REVIEW,
                    validation_status=ValidationStatus.MANUAL_REVIEW,
                    display_title=display or f"Satır {row_no}",
                    raw_data=raw,
                    normalized_data={},
                    validation_messages=[
                        {"code": "PARSED", "level": "INFO",
                         "msg": "Faz 3: parse edildi, modül validasyonu Faz 4'te eklenecek."}
                    ],
                )
                draft_count += 1
                total_rows += 1

            ImportLog.objects.create(
                batch=batch, level=ImportLogLevel.INFO, code="SHEET_PARSED",
                message=f"Sheet '{s_name}' parse edildi",
                context={"sheet": s_name, "rows": len(ws_rows), "drafts": draft_count},
            )

        return {
            "sheet_count": len(sheets),
            "total_rows": total_rows,
            "draft_count": draft_count,
            "skipped_empty": skipped_empty,
        }
    finally:
        wb.close()
