"""Rapor servisleri — Faz 12 MVP.

Senkron export (Celery YOK). XLSX (openpyxl) + CSV (UTF-8 BOM, ; ayraç).
AuditLog her export'ta yazılır.
Yetki: viewer export edemez; AUDIT raporu sadece manager/admin.
"""
from __future__ import annotations

import csv
import io
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Any, Dict, Iterable, List, Optional

from django.core.exceptions import PermissionDenied, ValidationError
from django.core.files.base import ContentFile
from django.db import transaction
from django.utils import timezone
from django.utils.text import slugify

from apps.audit.services import audit_log
from apps.finance.permissions import APPROVE_ROLES, WRITE_ROLES, can_approve, can_write

from ..models import (
    ReportFormat,
    ReportRun,
    ReportRunStatus,
    ReportTemplate,
    ReportType,
)


# ---------- Permission ----------


def _user_groups(user) -> set[str]:
    if not user or not getattr(user, "is_authenticated", False):
        return set()
    if user.is_superuser:
        return WRITE_ROLES | APPROVE_ROLES
    return set(user.groups.values_list("name", flat=True))


def can_run_report(user, report_type: str) -> bool:
    """Viewer hiçbir export yapamaz; AUDIT sadece manager/admin."""
    if not user or not getattr(user, "is_authenticated", False):
        return False
    if user.is_superuser:
        return True
    if not can_write(user):
        return False
    if report_type == ReportType.AUDIT:
        return can_approve(user)
    return True


# ---------- Definitions ----------


def _money(v) -> str:
    if v is None or v == "":
        return ""
    try:
        d = Decimal(str(v))
    except Exception:
        return str(v)
    s = f"{d:,.2f}"
    # 1,234,567.89 → 1.234.567,89
    s = s.replace(",", "X").replace(".", ",").replace("X", ".")
    return s


def _date(v) -> str:
    if not v:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%d.%m.%Y")
    return str(v)


def _datetime(v) -> str:
    if not v:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%d.%m.%Y %H:%M")
    return str(v)


# Each definition: title, columns [(key,label,formatter)], queryset_fn(filters,user), row_fn(obj)
REPORT_DEFINITIONS: Dict[str, dict] = {}


def _register(report_type: str, definition: dict):
    REPORT_DEFINITIONS[report_type] = definition


def _payable_qs(filters, user):
    from apps.finance.models import PayableItem, PayableStatus
    qs = PayableItem.objects.filter(is_active=True).select_related("institution")
    today = timezone.localdate()
    mode = filters.get("mode", "overdue")
    if mode == "overdue":
        qs = qs.filter(due_date__lt=today).exclude(
            status__in=[PayableStatus.PAID, PayableStatus.CANCELLED, PayableStatus.ARCHIVED]
        )
    elif mode == "upcoming":
        days = int(filters.get("days", 7))
        qs = qs.filter(due_date__gte=today, due_date__lte=today + timedelta(days=days)).exclude(
            status__in=[PayableStatus.PAID, PayableStatus.CANCELLED, PayableStatus.ARCHIVED]
        )
    elif mode == "missing_receipt":
        qs = qs.filter(requires_receipt=True)
    elif mode == "pending_approval":
        qs = qs.filter(status=PayableStatus.WAITING_APPROVAL)
    return qs.order_by("due_date")


def _payable_row(p):
    paid = getattr(p, "paid_amount", None) or Decimal("0")
    remain = (p.amount or Decimal("0")) - paid
    return {
        "owner": getattr(p, "owner_label", "") or "",
        "title": p.title,
        "category": getattr(p, "category", "") or "",
        "institution": str(p.institution) if getattr(p, "institution", None) else "",
        "period": getattr(p, "period_label", "") or "",
        "due_date": _date(p.due_date),
        "amount": _money(p.amount),
        "paid": _money(paid),
        "remaining": _money(remain),
        "status": p.get_status_display() if hasattr(p, "get_status_display") else "",
        "requires_receipt": "Evet" if getattr(p, "requires_receipt", False) else "Hayır",
    }


_PAYABLE_COLUMNS = [
    ("owner", "Sahip"),
    ("title", "Başlık"),
    ("category", "Kategori"),
    ("institution", "Kurum/Tedarikçi"),
    ("period", "Dönem"),
    ("due_date", "Son Ödeme"),
    ("amount", "Tutar"),
    ("paid", "Ödenen"),
    ("remaining", "Kalan"),
    ("status", "Durum"),
    ("requires_receipt", "Dekont Zorunlu"),
]


_register(ReportType.PAYABLES, {
    "title": "Geciken / Yaklaşan Ödemeler",
    "columns": _PAYABLE_COLUMNS,
    "queryset": _payable_qs,
    "row": _payable_row,
})


# Tasks
def _tasks_qs(filters, user):
    from apps.tasks.models import Task
    qs = Task.objects.filter(is_active=True).select_related("assigned_to", "created_by")
    if filters.get("status"):
        qs = qs.filter(status=filters["status"])
    if filters.get("assigned_to"):
        qs = qs.filter(assigned_to_id=filters["assigned_to"])
    if filters.get("due_from"):
        qs = qs.filter(due_date__gte=filters["due_from"])
    if filters.get("due_to"):
        qs = qs.filter(due_date__lte=filters["due_to"])
    return qs.order_by("due_date")


def _tasks_row(t):
    return {
        "title": t.title,
        "assigned_to": str(t.assigned_to) if t.assigned_to_id else "",
        "created_by": str(t.created_by) if t.created_by_id else "",
        "priority": t.get_priority_display(),
        "status": t.get_status_display(),
        "due_date": _date(t.due_date),
        "related": f"{t.related_app}/{t.related_model}#{t.related_object_id}" if t.related_app else "",
    }


_register(ReportType.TASKS, {
    "title": "Görev Raporu",
    "columns": [
        ("title", "Başlık"), ("assigned_to", "Atanan"), ("created_by", "Oluşturan"),
        ("priority", "Öncelik"), ("status", "Durum"), ("due_date", "Son Tarih"),
        ("related", "Bağlı Kayıt"),
    ],
    "queryset": _tasks_qs,
    "row": _tasks_row,
})


# Site Aidatları (PruvaStatement)
def _site_dues_qs(filters, user):
    from apps.pruva.models import PruvaStatement
    return PruvaStatement.objects.filter(is_active=True).select_related("unit").order_by("-year", "-month")


def _site_dues_row(s):
    return {
        "unit": str(s.unit) if getattr(s, "unit_id", None) else "",
        "year": s.year if hasattr(s, "year") else "",
        "month": s.month if hasattr(s, "month") else "",
        "due_date": _date(getattr(s, "due_date", None)),
        "total": _money(getattr(s, "total_amount", None) or getattr(s, "amount", None)),
        "status": s.get_status_display() if hasattr(s, "get_status_display") else "",
        "payable": str(getattr(s, "payable_id", "") or ""),
        "doc_status": "Var" if getattr(s, "document_id", None) else "Yok",
    }


_register(ReportType.SITE_DUES, {
    "title": "Site Aidatları Raporu",
    "columns": [
        ("unit", "Bağımsız Bölüm"), ("year", "Yıl"), ("month", "Ay"),
        ("due_date", "Son Ödeme"), ("total", "Toplam Tutar"),
        ("status", "Durum"), ("payable", "Payable"), ("doc_status", "Belge"),
    ],
    "queryset": _site_dues_qs,
    "row": _site_dues_row,
})


# Property tax
def _property_tax_qs(filters, user):
    from apps.properties.models import PropertyTaxInstallment
    return PropertyTaxInstallment.objects.filter(is_active=True).select_related("tax_year__property_asset").order_by("due_date")


def _property_tax_row(i):
    asset = getattr(i.tax_year, "property_asset", None) if getattr(i, "tax_year_id", None) else None
    return {
        "asset": str(asset) if asset else "",
        "year": getattr(i.tax_year, "year", "") if getattr(i, "tax_year_id", None) else "",
        "installment_no": getattr(i, "installment_no", ""),
        "municipality": str(getattr(asset, "municipality", "") or "") if asset else "",
        "due_date": _date(i.due_date),
        "amount": _money(i.amount),
        "status": i.get_status_display() if hasattr(i, "get_status_display") else "",
        "doc_status": "Var" if getattr(i, "document_id", None) else "Yok",
    }


_register(ReportType.PROPERTY_TAX, {
    "title": "Emlak Vergisi Raporu",
    "columns": [
        ("asset", "Mülk"), ("year", "Yıl"), ("installment_no", "Taksit"),
        ("municipality", "Belediye"), ("due_date", "Son Ödeme"),
        ("amount", "Tutar"), ("status", "Durum"), ("doc_status", "Belge"),
    ],
    "queryset": _property_tax_qs,
    "row": _property_tax_row,
})


# Guarantees commission
def _guarantee_qs(filters, user):
    from apps.guarantees.models import GuaranteeCommissionPeriod
    return GuaranteeCommissionPeriod.objects.filter(is_active=True).select_related("guarantee", "guarantee__bank").order_by("due_date")


def _guarantee_row(p):
    g = getattr(p, "guarantee", None)
    return {
        "letter_no": getattr(g, "letter_no", "") if g else "",
        "bank": str(getattr(g, "bank", "")) if g else "",
        "period": getattr(p, "period_label", "") or "",
        "due_date": _date(p.due_date),
        "amount": _money(getattr(p, "commission_amount", None) or getattr(p, "amount", None)),
        "status": p.get_status_display() if hasattr(p, "get_status_display") else "",
        "payable": str(getattr(p, "payable_id", "") or ""),
    }


_register(ReportType.GUARANTEES, {
    "title": "Teminat Komisyon Raporu",
    "columns": [
        ("letter_no", "Teminat"), ("bank", "Banka"), ("period", "Dönem"),
        ("due_date", "Son Ödeme"), ("amount", "Komisyon"), ("status", "Durum"),
        ("payable", "Payable"),
    ],
    "queryset": _guarantee_qs,
    "row": _guarantee_row,
})


# Integrators
def _integrators_qs(filters, user):
    from apps.integrators.models import CreditPackage
    return CreditPackage.objects.filter(is_active=True).select_related("service").order_by("-created_at")


def _integrators_row(pkg):
    s = getattr(pkg, "service", None)
    return {
        "service": str(s) if s else "",
        "provider": getattr(s, "provider_name", "") if s else "",
        "package": getattr(pkg, "package_name", "") or str(pkg),
        "remaining": getattr(pkg, "remaining_credits", "") or "",
        "threshold": getattr(pkg, "critical_threshold", "") or "",
        "status": pkg.get_status_display() if hasattr(pkg, "get_status_display") else "",
    }


_register(ReportType.INTEGRATORS, {
    "title": "Entegratör / Kontör Raporu",
    "columns": [
        ("service", "Hizmet"), ("provider", "Sağlayıcı"), ("package", "Paket"),
        ("remaining", "Kalan Kontör"), ("threshold", "Kritik Eşik"), ("status", "Durum"),
    ],
    "queryset": _integrators_qs,
    "row": _integrators_row,
})


# Documents
def _docs_qs(filters, user):
    from apps.documents.models import Document
    return Document.objects.filter(is_active=True).select_related("uploaded_by").order_by("-created_at")


def _docs_row(d):
    return {
        "title": d.title or (d.file.name if getattr(d, "file", None) else ""),
        "type": d.get_document_type_display() if hasattr(d, "get_document_type_display") else "",
        "size": getattr(d, "file_size", "") or "",
        "sha_short": (d.sha256[:12] if getattr(d, "sha256", None) else ""),
        "uploaded_by": str(d.uploaded_by) if d.uploaded_by_id else "",
        "uploaded_at": _datetime(d.created_at),
    }


_register(ReportType.DOCUMENTS, {
    "title": "Belge Raporu",
    "columns": [
        ("title", "Dosya Adı"), ("type", "Tür"), ("size", "Boyut"),
        ("sha_short", "SHA-256"), ("uploaded_by", "Yükleyen"), ("uploaded_at", "Tarih"),
    ],
    "queryset": _docs_qs,
    "row": _docs_row,
})


# Audit
def _audit_qs(filters, user):
    from apps.audit.models import AuditLog
    return AuditLog.objects.select_related("actor").order_by("-created_at")


def _audit_row(a):
    return {
        "ts": _datetime(a.created_at),
        "user": str(a.actor) if a.actor_id else "",
        "action": a.get_action_display() if hasattr(a, "get_action_display") else a.action,
        "model": f"{a.app_label}/{a.model_name}",
        "obj": a.object_repr or a.object_id,
        "summary": a.summary,
        "ip": a.ip_address or "",
    }


_register(ReportType.AUDIT, {
    "title": "AuditLog Raporu",
    "columns": [
        ("ts", "Zaman"), ("user", "Kullanıcı"), ("action", "Aksiyon"),
        ("model", "Model"), ("obj", "Obje"), ("summary", "Özet"), ("ip", "IP"),
    ],
    "queryset": _audit_qs,
    "row": _audit_row,
})


ALL_REPORT_TYPES = list(REPORT_DEFINITIONS.keys())


# ---------- Public API ----------


def get_report_definition(report_type: str) -> dict:
    if report_type not in REPORT_DEFINITIONS:
        raise ValidationError(f"Bilinmeyen rapor tipi: {report_type}")
    return REPORT_DEFINITIONS[report_type]


def get_available_report_types(user) -> List[dict]:
    items = []
    for rt in ALL_REPORT_TYPES:
        if can_run_report(user, rt):
            d = REPORT_DEFINITIONS[rt]
            items.append({"key": rt, "title": d["title"]})
    return items


def build_report_queryset(report_type: str, filters: dict, user):
    d = get_report_definition(report_type)
    return d["queryset"](filters or {}, user)


def apply_common_filters(qs, filters: dict):
    """Genel is_active / created_at filtreleri."""
    if not filters:
        return qs
    if filters.get("created_from"):
        qs = qs.filter(created_at__gte=filters["created_from"])
    if filters.get("created_to"):
        qs = qs.filter(created_at__lte=filters["created_to"])
    return qs


def build_report_rows(report_type: str, qs, columns: Optional[List[str]] = None) -> List[dict]:
    d = get_report_definition(report_type)
    rows = []
    for obj in qs:
        try:
            row = d["row"](obj)
        except Exception as exc:  # noqa: BLE001
            row = {"_error": str(exc)}
        rows.append(row)
    return rows


def preview_report(report_type: str, filters: dict, user, *, limit: int = 50) -> dict:
    if not can_run_report(user, report_type):
        raise PermissionDenied("Bu raporu çalıştırma yetkiniz yok.")
    d = get_report_definition(report_type)
    qs = build_report_queryset(report_type, filters or {}, user)
    total = qs.count()
    rows = build_report_rows(report_type, list(qs[:limit]))
    return {
        "report_type": report_type,
        "title": d["title"],
        "columns": d["columns"],
        "rows": rows,
        "row_count": total,
        "preview_count": len(rows),
    }


# ---------- Export ----------


def _safe_filename(report_type: str, fmt: str) -> str:
    ts = timezone.now().strftime("%Y%m%d_%H%M%S")
    return f"{slugify(report_type.lower())}_{ts}.{fmt.lower()}"


def export_to_xlsx(rows, columns, *, title: str, metadata: dict | None = None) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    sheet_name = (title or "Rapor")[:31].replace(":", "-").replace("/", "-").replace("\\", "-")
    ws.title = sheet_name or "Rapor"

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E5E7EB")

    # Header rows
    ws.append([title or "Rapor"])
    ws["A1"].font = Font(bold=True, size=14)
    meta = metadata or {}
    info_line = f"Üreten: {meta.get('actor','')} · Üretildi: {timezone.now().strftime('%d.%m.%Y %H:%M')}"
    ws.append([info_line])
    ws.append([])  # blank row

    # Column header
    headers = [label for _, label in columns]
    ws.append(headers)
    header_row_idx = ws.max_row
    for cell in ws[header_row_idx]:
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    keys = [k for k, _ in columns]
    for r in rows:
        ws.append([r.get(k, "") for k in keys])

    # Auto-width
    for col_idx, _ in enumerate(columns, start=1):
        col_letter = ws.cell(row=header_row_idx, column=col_idx).column_letter
        max_len = max(
            [len(str(ws.cell(row=row_i, column=col_idx).value or "")) for row_i in range(header_row_idx, ws.max_row + 1)]
            + [12]
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def export_to_csv(rows, columns, *, title: str, metadata: dict | None = None) -> bytes:
    out = io.StringIO()
    out.write("\ufeff")  # UTF-8 BOM
    writer = csv.writer(out, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    writer.writerow([title])
    meta = metadata or {}
    writer.writerow([f"Üreten: {meta.get('actor','')}"])
    writer.writerow([f"Üretildi: {timezone.now().strftime('%d.%m.%Y %H:%M')}"])
    writer.writerow([])
    writer.writerow([label for _, label in columns])
    keys = [k for k, _ in columns]
    for r in rows:
        writer.writerow([r.get(k, "") for k in keys])
    return out.getvalue().encode("utf-8")


@transaction.atomic
def run_export(
    *,
    report_type: str,
    filters: dict,
    columns: Optional[list] = None,
    fmt: str = ReportFormat.XLSX,
    user,
    template: Optional[ReportTemplate] = None,
    title: str = "",
    request=None,
) -> ReportRun:
    if not can_run_report(user, report_type):
        raise PermissionDenied("Export yetkiniz yok.")

    run = ReportRun.objects.create(
        template=template,
        report_type=report_type,
        title=title or report_type,
        filters=filters or {},
        columns=columns or [],
        format=fmt,
        status=ReportRunStatus.RUNNING,
        created_by=user if getattr(user, "is_authenticated", False) else None,
    )
    try:
        d = get_report_definition(report_type)
        title = title or d["title"]
        if not run.columns:
            run.columns = [c[0] for c in d["columns"]]
        run.title = title
        qs = build_report_queryset(report_type, filters or {}, user)
        rows = build_report_rows(report_type, list(qs))
        meta = {"actor": str(user) if user else ""}
        if fmt == ReportFormat.XLSX:
            content = export_to_xlsx(rows, d["columns"], title=title, metadata=meta)
        elif fmt == ReportFormat.CSV:
            content = export_to_csv(rows, d["columns"], title=title, metadata=meta)
        else:
            raise ValidationError(f"Bilinmeyen format: {fmt}")
        fname = _safe_filename(report_type, fmt)
        run.output_file.save(fname, ContentFile(content), save=False)
        run.row_count = len(rows)
        run.file_size = len(content)
        run.status = ReportRunStatus.COMPLETED
        run.completed_at = timezone.now()
        run.save()
        audit_log(
            actor=user, action="EXPORT", obj=run,
            summary=f"Rapor export: {report_type} · {fmt} · {len(rows)} satır",
            metadata={"report_type": report_type, "format": fmt, "rows": len(rows)},
            request=request,
        )
        return run
    except Exception as exc:  # noqa: BLE001
        run.status = ReportRunStatus.FAILED
        run.error_message = str(exc)[:1000]
        run.completed_at = timezone.now()
        run.save(update_fields=["status", "error_message", "completed_at", "updated_at"])
        audit_log(
            actor=user, action="EXPORT", obj=run,
            summary=f"Rapor export FAILED: {report_type}",
            metadata={"error": str(exc)[:200]}, request=request,
        )
        return run


# ---------- Template management ----------


@transaction.atomic
def create_report_template(
    *, name: str, slug: str, report_type: str, user,
    description: str = "", default_format: str = ReportFormat.XLSX,
    default_filters: Optional[dict] = None, columns: Optional[list] = None,
    request=None,
) -> ReportTemplate:
    if not can_write(user):
        raise PermissionDenied("Şablon oluşturma yetkisi yok.")
    if report_type not in REPORT_DEFINITIONS:
        raise ValidationError(f"Bilinmeyen rapor tipi: {report_type}")
    t = ReportTemplate.objects.create(
        name=name.strip(),
        slug=slug.strip() or slugify(name),
        report_type=report_type,
        description=description or "",
        default_format=default_format,
        default_filters=default_filters or {},
        columns=columns or [c[0] for c in REPORT_DEFINITIONS[report_type]["columns"]],
        created_by=user if getattr(user, "is_authenticated", False) else None,
    )
    audit_log(actor=user, action="CREATE", obj=t,
              summary=f"Rapor şablonu: {t}", request=request)
    return t


@transaction.atomic
def update_report_template(
    *, template: ReportTemplate, user, request=None, **fields
) -> ReportTemplate:
    if not can_write(user):
        raise PermissionDenied("Şablon güncelleme yetkisi yok.")
    for k, v in fields.items():
        if hasattr(template, k):
            setattr(template, k, v)
    template.save()
    audit_log(actor=user, action="UPDATE", obj=template,
              summary=f"Şablon güncellendi: {template}", request=request)
    return template


@transaction.atomic
def cancel_report_run(*, run: ReportRun, user, request=None) -> ReportRun:
    if run.status not in (ReportRunStatus.PENDING, ReportRunStatus.RUNNING):
        return run
    if not (user and (user.is_superuser or run.created_by_id == user.id or can_approve(user))):
        raise PermissionDenied("İptal yetkisi yok.")
    run.status = ReportRunStatus.CANCELLED
    run.save(update_fields=["status", "updated_at"])
    audit_log(actor=user, action="UPDATE", obj=run,
              summary="Rapor iptal edildi", request=request)
    return run


def get_recent_exports(user, *, limit: int = 10):
    qs = ReportRun.objects.filter(is_active=True)
    if not (user and user.is_superuser):
        if not can_approve(user):
            qs = qs.filter(created_by=user)
    return qs.order_by("-created_at")[:limit]
